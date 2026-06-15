"""Read raw rows from an MTS brokerage report (.xlsx)."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from package_snowball.core.models import (
    CashMovementRow,
    DealCompletedRow,
    SecurityBalanceRow,
)


def _to_decimal(value: object) -> Decimal:
    """Convert a cell value to Decimal, treating None/empty as 0."""
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    return Decimal(str(value).replace(" ", "").replace(",", "."))


def _to_str(value: object) -> str:
    """Convert a cell value to string, treating None as empty."""
    if value is None:
        return ""
    return str(value).strip()


class MtsReportReader:
    """Parse an MTS .xlsx report into raw row dataclasses."""

    def __init__(self, path: Path) -> None:
        self.path = path

    def read(
        self,
    ) -> tuple[list[DealCompletedRow], list[CashMovementRow], list[SecurityBalanceRow]]:
        """Return (deals_completed, cash_movements, securities_balances)."""
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Workbook has no active sheet")

        deals = self._read_deals_completed(ws)
        cash = self._read_cash_movements(ws)
        balances = self._read_securities_movement(ws)
        return deals, cash, balances

    def _read_deals_completed(self, ws: Worksheet) -> list[DealCompletedRow]:
        """Read the 'deals completed' section."""
        start_row = self._find_row(ws, "Информация о сделках, завершенных в отчетном периоде:")
        if start_row is None:
            return []

        # Skip header row and column-number row
        data_start = start_row + 2
        result: list[DealCompletedRow] = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                break
            # Skip the column-number row ("1", "2", "3"...)
            if _to_str(row[0]) == "1" and _to_str(row[1]) == "2":
                continue
            result.append(
                DealCompletedRow(
                    deal_number=_to_str(row[0]),
                    deal_date=_to_str(row[1]),
                    deal_type=_to_str(row[2]),
                    counterparty=_to_str(row[3]),
                    security_name=_to_str(row[4]),
                    isin=_to_str(row[5]),
                    quantity=_to_decimal(row[6]),
                    price=_to_decimal(row[7]),
                    currency=_to_str(row[8]),
                    amount_without_nkd=_to_decimal(row[9]),
                    nkd=_to_decimal(row[10]),
                    deal_amount=_to_decimal(row[11]),
                    payment_currency=_to_str(row[12]),
                    trading_system=_to_str(row[13]),
                    delivery_date=_to_str(row[14]),
                    payment_date=_to_str(row[15]),
                    exchange_commission=_to_decimal(row[16]),
                    exchange_commission_currency=_to_str(row[17]),
                    broker_commission=_to_decimal(row[18]),
                    broker_commission_currency=_to_str(row[19]),
                )
            )
        return result

    def _read_cash_movements(self, ws: Worksheet) -> list[CashMovementRow]:
        """Read the 'cash movement' section."""
        section_row = self._find_row(ws, "Информация о движении денежных средств")
        if section_row is None:
            return []

        result: list[CashMovementRow] = []
        current_currency = ""

        for row in ws.iter_rows(min_row=section_row + 1, values_only=True):
            first_col = _to_str(row[0])

            # Detect currency sub-section header, e.g. "Входящий остаток ... (RUB):"
            if first_col.startswith("Входящий остаток денежных средств ("):
                current_currency = self._extract_currency(first_col)
                continue

            # Stop at totals, empty rows, or footer rows
            if first_col == "" or first_col.startswith("Итого"):
                current_currency = ""
                continue
            if first_col.startswith("Исходящий остаток") or first_col.startswith(
                "Плановый остаток"
            ):
                continue
            if first_col.startswith("в том числе"):
                continue

            # Header rows inside the cash section
            if first_col in ("Дата", "1", "2"):
                continue

            if current_currency and first_col:
                result.append(
                    CashMovementRow(
                        date=_to_str(row[0]),
                        operation=_to_str(row[1]),
                        credited=_to_decimal(row[7]),
                        debited=_to_decimal(row[9]),
                        currency=current_currency,
                    )
                )

        return result

    def _read_securities_movement(self, ws: Worksheet) -> list[SecurityBalanceRow]:
        """Read the 'securities movement' section."""
        start_row = self._find_row(ws, "Информация о движении ценных бумаг")
        if start_row is None:
            return []

        # Skip header row and column-number row
        data_start = start_row + 2
        result: list[SecurityBalanceRow] = []
        for row in ws.iter_rows(min_row=data_start, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                break
            # Skip the column-number row ("1", "2", "3"...)
            if _to_str(row[0]) == "1" and _to_str(row[1]) == "2":
                continue
            result.append(
                SecurityBalanceRow(
                    security_name=_to_str(row[0]),
                    isin=_to_str(row[1]),
                    opening_balance=_to_decimal(row[2]),
                )
            )
        return result

    @staticmethod
    def _find_row(ws: Worksheet, text: str) -> int | None:
        """Return the 1-based row number where the first column matches text."""
        for idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
        ):
            if row and row[0] == text:
                return idx
        return None

    @staticmethod
    def _extract_currency(header: str) -> str:
        """Extract currency code from a header like 'Входящий остаток ... (RUB):'."""
        start = header.find("(")
        end = header.find(")")
        if start != -1 and end != -1:
            return header[start + 1 : end]
        return ""
